from __future__ import annotations

import datetime as dt
import logging
import shutil
from pathlib import Path

from openpyxl import load_workbook

from filldoc.core.errors import ExcelError
from .models import Project

log = logging.getLogger("filldoc.excel")


class ExcelProjectStore:
    """
    - читаем лист «Текущие» (иначе первый лист)
    - первая строка = заголовки столбцов
    - проект = строка, где есть хоть одно значение
    - идентификатор проекта: колонка с номером дела (несколько вариантов названия), иначе 'row:<номер>'
    """

    # Возможные названия колонки с номером дела в Excel.
    # Поддерживаем старое и новое имя (после переименования в UI).
    _CASE_NUMBER_HEADERS: tuple[str, ...] = (
        "№ дела",
        "№дела",
        "Номер дела",
        "номер дела",
        "Номер осн. дела",
        "номер осн. дела",
        "Номер осн дела",
        "номер осн дела",
    )

    def __init__(self, excel_path: str) -> None:
        self.excel_path = excel_path

    # ------------------------------------------------------------------ публичные методы загрузки

    def load_projects(self) -> list[Project]:
        path = Path(self.excel_path)
        if not path.exists():
            raise ExcelError("Excel-файл недоступен или не существует.")
        try:
            wb = load_workbook(filename=path, read_only=False, data_only=True)
            ws = wb["Текущие"] if "Текущие" in wb.sheetnames else wb.worksheets[0]
            return self._load_projects_from_worksheet(ws)
        except ExcelError:
            raise
        except Exception as e:  # noqa: BLE001
            raise ExcelError(f"Не удалось загрузить проекты из Excel: {e}") from e

    def load_projects_from_sheet(self, sheet_name: str) -> list[Project]:
        """Загружает проекты с указанного листа; если лист не найден — возвращает []."""
        path = Path(self.excel_path)
        if not path.exists():
            raise ExcelError("Excel-файл недоступен или не существует.")
        try:
            wb = load_workbook(filename=path, read_only=False, data_only=True)
            ws = self._find_sheet(wb, sheet_name)
            if ws is None:
                return []
            return self._load_projects_from_worksheet(ws)
        except ExcelError:
            raise
        except Exception as e:  # noqa: BLE001
            raise ExcelError(f"Не удалось загрузить проекты из листа '{sheet_name}': {e}") from e

    def repair_archive_headers(self, archive_sheet_name: str = "Архив") -> bool:
        """
        Проверяет лист архива: если первая строка не совпадает с заголовками листа
        «Текущие», вставляет правильные заголовки в строку 1.
        Возвращает True, если была выполнена починка, False если всё было в порядке.
        """
        path = Path(self.excel_path)
        if not path.exists():
            raise ExcelError("Excel-файл недоступен или не существует.")
        try:
            wb = load_workbook(filename=path, read_only=False, data_only=False)
            ws_archive = self._find_sheet(wb, archive_sheet_name)
            if ws_archive is None:
                return False

            ws_current = wb["Текущие"] if "Текущие" in wb.sheetnames else wb.worksheets[0]
            expected_headers = [
                str(c.value).strip() if c.value is not None else ""
                for c in ws_current[1]
            ]

            archive_row1 = [
                str(c.value).strip() if c.value is not None else ""
                for c in ws_archive[1]
            ] if (ws_archive.max_row or 0) >= 1 else []

            # Если первая строка архива уже совпадает с заголовками — всё в порядке.
            if archive_row1 == expected_headers:
                return False

            # Если у листа вообще нет строк — просто добавляем заголовки.
            if (ws_archive.max_row or 0) < 1:
                ws_archive.append(expected_headers)
                wb.save(path)
                return True

            # Если первая строка похожа на старую шапку, обновляем её на месте,
            # а не вставляем новую: иначе в данных остаётся лишняя строка «Имя проекта».
            if self._is_header_like_row(archive_row1, expected_headers):
                for col_idx, header_val in enumerate(expected_headers, start=1):
                    ws_archive.cell(row=1, column=col_idx).value = header_val
            else:
                ws_archive.insert_rows(1)
                for col_idx, header_val in enumerate(expected_headers, start=1):
                    ws_archive.cell(row=1, column=col_idx).value = header_val
            wb.save(path)
            return True
        except ExcelError:
            raise
        except Exception as e:  # noqa: BLE001
            raise ExcelError(f"Не удалось восстановить заголовки архива: {e}") from e

    # ------------------------------------------------------------------ создание бэкапа

    _BACKUP_MAX_COUNT: int = 20
    _BACKUP_MAX_DAYS: int = 15

    def create_backup(self) -> Path:
        src = Path(self.excel_path)
        if not src.exists():
            raise ExcelError("Не удалось создать резервную копию: Excel-файл не найден.")
        backup_dir = src.parent / "_filldoc_backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        ts = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
        dst = backup_dir / f"{src.stem}__backup__{ts}{src.suffix}"
        try:
            shutil.copy2(src, dst)
        except PermissionError as e:
            raise ExcelError(
                "Нет доступа к Excel-файлу. Закройте Excel (если файл открыт), "
                "снимите атрибут 'Только чтение' и проверьте права на папку."
            ) from e
        except Exception as e:  # noqa: BLE001
            raise ExcelError(f"Не удалось создать резервную копию Excel-файла: {e}") from e

        log.debug("Backup created: %s", dst)
        self._rotate_backups(backup_dir, src.stem, src.suffix)
        return dst

    def _rotate_backups(self, backup_dir: Path, stem: str, suffix: str) -> None:
        """Удаляет бэкапы старше _BACKUP_MAX_DAYS дней и сверх _BACKUP_MAX_COUNT штук."""
        pattern = f"{stem}__backup__*{suffix}"
        backups = sorted(backup_dir.glob(pattern), key=lambda p: p.stat().st_mtime)

        cutoff = dt.datetime.now() - dt.timedelta(days=self._BACKUP_MAX_DAYS)
        to_delete: list[Path] = []

        for bp in backups:
            try:
                mtime = dt.datetime.fromtimestamp(bp.stat().st_mtime)
                if mtime < cutoff:
                    to_delete.append(bp)
            except OSError:
                pass

        # Удаляем по возрасту
        for bp in to_delete:
            try:
                bp.unlink()
            except OSError:
                pass

        # После удаления устаревших — оставляем не более _BACKUP_MAX_COUNT
        remaining = sorted(
            (p for p in backup_dir.glob(pattern) if p not in to_delete),
            key=lambda p: p.stat().st_mtime,
        )
        excess = len(remaining) - self._BACKUP_MAX_COUNT
        for bp in remaining[:max(excess, 0)]:
            try:
                bp.unlink()
            except OSError:
                pass

    # ------------------------------------------------------------------ сохранение / добавление

    def save_project_fields(self, project: Project) -> None:
        """Записывает поля проекта обратно в Excel по row_index."""
        path = Path(self.excel_path)
        if not path.exists():
            raise ExcelError("Excel-файл недоступен или не существует.")

        self.create_backup()

        try:
            wb = load_workbook(filename=path, read_only=False, data_only=False)
            ws = wb.worksheets[0]
            headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
            target_row = project.row_index
            if target_row is None or target_row < 2 or target_row > (ws.max_row or 0):
                raise ExcelError("Не удалось определить строку проекта в Excel для сохранения.")

            header_to_col = {headers[i]: i + 1 for i in range(len(headers)) if headers[i]}
            for k, v in project.fields.items():
                if k in header_to_col:
                    ws.cell(row=target_row, column=header_to_col[k]).value = v
            wb.save(path)
        except PermissionError as e:
            raise ExcelError(
                "Не удалось сохранить изменения: Excel-файл занят или недоступен для записи. "
                "Закройте Excel (если файл открыт) и попробуйте снова."
            ) from e
        except ExcelError:
            raise
        except Exception as e:  # noqa: BLE001
            raise ExcelError(f"Не удалось сохранить изменения в Excel: {e}") from e

    def add_project(self, project: Project) -> None:
        """Добавляет новую строку в конец Excel-листа и обновляет project.row_index."""
        path = Path(self.excel_path)
        if not path.exists():
            raise ExcelError("Excel-файл недоступен или не существует.")

        self.create_backup()

        try:
            wb = load_workbook(filename=path, read_only=False, data_only=False)
            ws = wb.worksheets[0]
            headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]

            header_to_idx = {h: i for i, h in enumerate(headers) if h}
            new_row: list[str] = [""] * len(headers)
            for k, v in project.fields.items():
                if k in header_to_idx:
                    new_row[header_to_idx[k]] = v

            ws.append(new_row)
            project.row_index = ws.max_row
            project.headers = [h for h in headers if h]
            wb.save(path)
        except PermissionError as e:
            raise ExcelError(
                "Не удалось добавить проект: Excel-файл занят или недоступен для записи. "
                "Закройте Excel (если файл открыт) и попробуйте снова."
            ) from e
        except ExcelError:
            raise
        except Exception as e:  # noqa: BLE001
            raise ExcelError(f"Не удалось добавить проект в Excel: {e}") from e

    def save_all_projects(
        self,
        active_projects: list[Project],
        archived_projects: list[Project] | None = None,
    ) -> None:
        """
        Полная синхронизация проектов с Excel:
        - лист «Текущие» (или первый): текущие проекты
        - лист «Архив» (опционально): архивные проекты
        """
        path = Path(self.excel_path)
        if not path.exists():
            raise ExcelError("Excel-файл недоступен или не существует.")

        self.create_backup()

        try:
            wb = load_workbook(filename=path, read_only=False, data_only=False)

            ws_current = wb["Текущие"] if "Текущие" in wb.sheetnames else wb.worksheets[0]
            headers = [str(c.value).strip() if c.value is not None else "" for c in ws_current[1]]
            if not any(headers):
                raise ExcelError("Не найдена строка заголовков в Excel.")

            if (ws_current.max_row or 0) > 1:
                ws_current.delete_rows(2, (ws_current.max_row or 1) - 1)

            header_to_idx = {h: i for i, h in enumerate(headers) if h}

            for idx, project in enumerate(active_projects, start=2):
                row_values: list[str] = [""] * len(headers)
                for k, v in project.fields.items():
                    if k in header_to_idx:
                        row_values[header_to_idx[k]] = v
                ws_current.append(row_values)
                project.row_index = idx
                project.headers = [h for h in headers if h]

            if archived_projects is not None:
                ws_archive = self._find_sheet(wb, "Архив")
                if ws_archive is None:
                    ws_archive = wb.create_sheet(title="Архив")
                    ws_archive.append(headers)
                else:
                    self._ensure_headers_on_sheet(ws_archive, headers)

                if (ws_archive.max_row or 0) > 1:
                    ws_archive.delete_rows(2, (ws_archive.max_row or 1) - 1)

                for idx, project in enumerate(archived_projects, start=2):
                    row_values = [""] * len(headers)
                    for k, v in project.fields.items():
                        if k in header_to_idx:
                            row_values[header_to_idx[k]] = v
                    ws_archive.append(row_values)
                    project.row_index = idx
                    project.headers = [h for h in headers if h]

            wb.save(path)
        except PermissionError as e:
            raise ExcelError(
                "Не удалось синхронизировать проекты: Excel-файл занят или недоступен для записи. "
                "Закройте Excel (если файл открыт) и попробуйте снова."
            ) from e
        except ExcelError:
            raise
        except Exception as e:  # noqa: BLE001
            raise ExcelError(f"Не удалось синхронизировать проекты с Excel: {e}") from e

    def move_project_to_archive(
        self,
        project: Project,
        current_sheet_name: str = "Текущие",
        archive_sheet_name: str = "Архив",
    ) -> None:
        """Переносит строку проекта с листа текущих проектов на лист архива."""
        path = Path(self.excel_path)
        if not path.exists():
            raise ExcelError("Excel-файл недоступен или не существует.")

        case_number = self._extract_case_number(project.fields)
        if not case_number and project.row_index is None:
            raise ExcelError(
                "Невозможно архивировать проект: не найдено поле с номером дела "
                "и не задан номер строки в Excel."
            )

        self.create_backup()

        try:
            wb = load_workbook(filename=path, read_only=False, data_only=False)

            ws_current = (
                wb[current_sheet_name]
                if current_sheet_name in wb.sheetnames
                else wb.worksheets[0]
            )

            ws_archive = self._find_sheet(wb, archive_sheet_name)
            if ws_archive is None:
                ws_archive = wb.create_sheet(title=archive_sheet_name)

            # Гарантируем наличие корректных заголовков на листе архива.
            current_headers = [c.value for c in ws_current[1]]
            self._ensure_headers_on_sheet(ws_archive, current_headers)

            row_idx = self._find_row_by_case_number(ws_current, case_number) if case_number else None
            if row_idx is None:
                row_idx = project.row_index
            if row_idx is None or row_idx < 2 or row_idx > (ws_current.max_row or 0):
                raise ExcelError("Не удалось определить строку проекта в Excel для архивации.")

            row_values = [cell.value for cell in ws_current[row_idx]]
            ws_archive.append(row_values)
            project.row_index = ws_archive.max_row

            ws_current.delete_rows(row_idx, 1)
            wb.save(path)
        except PermissionError as e:
            raise ExcelError(
                "Не удалось перенести проект в архив: Excel-файл занят или недоступен для записи. "
                "Закройте Excel (если файл открыт) и попробуйте снова."
            ) from e
        except ExcelError:
            raise
        except Exception as e:  # noqa: BLE001
            raise ExcelError(f"Не удалось перенести проект в архив: {e}") from e

    def restore_project_from_archive(
        self,
        project: Project,
        current_sheet_name: str = "Текущие",
        archive_sheet_name: str = "Архив",
    ) -> None:
        """Переносит строку проекта с листа архива обратно на лист текущих проектов."""
        path = Path(self.excel_path)
        if not path.exists():
            raise ExcelError("Excel-файл недоступен или не существует.")

        self.create_backup()

        try:
            wb = load_workbook(filename=path, read_only=False, data_only=False)

            ws_archive = self._find_sheet(wb, archive_sheet_name)
            if ws_archive is None:
                raise ExcelError("Лист архива в Excel не найден.")

            ws_current = (
                wb[current_sheet_name]
                if current_sheet_name in wb.sheetnames
                else wb.worksheets[0]
            )

            row_idx = self._resolve_archive_row(ws_archive, project)
            if row_idx is None:
                raise ExcelError(
                    "Не удалось найти строку проекта в архиве.\n"
                    "Возможно, файл был изменён внешней программой."
                )

            row_values = [cell.value for cell in ws_archive[row_idx]]
            ws_current.append(row_values)
            project.row_index = ws_current.max_row

            ws_archive.delete_rows(row_idx, 1)
            wb.save(path)
        except PermissionError as e:
            raise ExcelError(
                "Не удалось вернуть проект из архива: Excel-файл занят или недоступен для записи. "
                "Закройте Excel (если файл открыт) и попробуйте снова."
            ) from e
        except ExcelError:
            raise
        except Exception as e:  # noqa: BLE001
            raise ExcelError(f"Не удалось вернуть проект из архива: {e}") from e

    def delete_project(self, project: Project) -> None:
        """Удаляет строку проекта из Excel по row_index."""
        path = Path(self.excel_path)
        if not path.exists():
            raise ExcelError("Excel-файл недоступен или не существует.")

        if project.row_index is None:
            raise ExcelError("Невозможно удалить проект: не задан номер строки в Excel.")

        self.create_backup()

        try:
            wb = load_workbook(filename=path, read_only=False, data_only=False)
            ws = wb.worksheets[0]
            target_row = project.row_index
            if target_row < 2 or target_row > (ws.max_row or 0):
                raise ExcelError("Не удалось определить строку проекта в Excel для удаления.")
            ws.delete_rows(target_row, 1)
            wb.save(path)
        except PermissionError as e:
            raise ExcelError(
                "Не удалось удалить проект: Excel-файл занят или недоступен для записи. "
                "Закройте Excel (если файл открыт) и попробуйте снова."
            ) from e
        except ExcelError:
            raise
        except Exception as e:  # noqa: BLE001
            raise ExcelError(f"Не удалось удалить проект из Excel: {e}") from e

    def delete_project_from_archive(
        self,
        project: Project,
        archive_sheet_name: str = "Архив",
    ) -> None:
        """Удаляет строку проекта из листа архива Excel."""
        path = Path(self.excel_path)
        if not path.exists():
            raise ExcelError("Excel-файл недоступен или не существует.")

        self.create_backup()

        try:
            wb = load_workbook(filename=path, read_only=False, data_only=False)
            ws_archive = self._find_sheet(wb, archive_sheet_name)
            if ws_archive is None:
                raise ExcelError("Лист архива в Excel не найден.")

            row_idx = self._resolve_archive_row(ws_archive, project)
            if row_idx is None:
                raise ExcelError(
                    "Не удалось найти строку проекта в архиве.\n"
                    "Возможно, файл был изменён внешней программой."
                )

            ws_archive.delete_rows(row_idx, 1)
            wb.save(path)
        except PermissionError as e:
            raise ExcelError(
                "Не удалось удалить проект из архива: Excel-файл занят или недоступен для записи. "
                "Закройте Excel (если файл открыт) и попробуйте снова."
            ) from e
        except ExcelError:
            raise
        except Exception as e:  # noqa: BLE001
            raise ExcelError(f"Не удалось удалить проект из архива: {e}") from e

    # ------------------------------------------------------------------ приватные вспомогательные

    def _load_projects_from_worksheet(self, ws) -> list[Project]:  # noqa: ANN001
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        projects: list[Project] = []
        for idx, row in enumerate(rows[1:], start=2):
            if row is None:
                continue
            values = ["" if v is None else str(v) for v in row]
            if all(v.strip() == "" for v in values):
                continue
            if self._is_header_like_row(values, headers):
                continue
            fields = {
                headers[i]: values[i]
                for i in range(min(len(headers), len(values)))
                if headers[i]
            }
            pid = self._extract_case_number(fields) or f"row:{idx}"
            projects.append(
                Project(
                    project_id=pid,
                    fields=fields,
                    headers=headers,
                    row_index=idx,
                )
            )
        return projects

    def _is_header_like_row(self, values: list[str], headers: list[str]) -> bool:
        """Пропускает случайно продублированные строки заголовков внутри данных.

        Такое бывает, когда шапка листа обновляется/восстанавливается и старая
        строка заголовков остаётся ниже первой строки. Без этой проверки Excel-
        загрузчик воспринимает её как проект с названием «Имя проекта».
        """
        normalized_values = [str(v).strip() for v in values]
        non_empty_values = [v for v in normalized_values if v]
        if len(non_empty_values) < 2:
            return False

        header_set = {h for h in headers if h}
        if not header_set:
            return False

        if any(v not in header_set for v in non_empty_values):
            return False

        positional_matches = sum(
            1
            for i, value in enumerate(normalized_values[: len(headers)])
            if value and value == headers[i]
        )
        return positional_matches >= max(2, len(non_empty_values) // 2)

    def _find_sheet(self, wb, sheet_name: str):  # noqa: ANN001
        """Ищет лист сначала точно, затем без учёта регистра/пробелов. Возвращает None если не найден."""
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        target = sheet_name.strip().lower()
        for name in wb.sheetnames:
            if name.strip().lower() == target:
                return wb[name]
        return None

    def _ensure_headers_on_sheet(self, ws, headers: list) -> None:  # noqa: ANN001
        """
        Гарантирует, что первая строка листа ws содержит именно переданные заголовки.
        - Если лист пуст — добавляет строку заголовков.
        - Если первая строка уже совпадает — ничего не делает.
        - Если первая строка не совпадает (данные вместо заголовков) — вставляет заголовки перед ней.
        """
        max_r = ws.max_row or 0
        if max_r < 1:
            ws.append(headers)
            return

        existing_row1 = [
            str(c.value).strip() if c.value is not None else ""
            for c in ws[1]
        ]
        expected = [str(h).strip() if h is not None else "" for h in headers]

        if max_r == 1 and not any(existing_row1):
            for col_idx, header_val in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx).value = header_val
            return

        if existing_row1 == expected:
            return  # уже всё правильно

        # Если первая строка похожа на устаревшую шапку, обновляем её на месте.
        if self._is_header_like_row(existing_row1, expected):
            for col_idx, header_val in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx).value = header_val
            return

        # Первая строка содержит данные — вставляем заголовки перед ней.
        ws.insert_rows(1)
        for col_idx, header_val in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx).value = header_val

    # Значения-заглушки, которые не считаются настоящим номером дела.
    _PLACEHOLDER_VALUES: frozenset[str] = frozenset({"—", "-", "–", "−", "n/a", "нет", "."})

    def _extract_case_number(self, fields: dict[str, str]) -> str:
        """Возвращает номер дела из словаря полей по любому из допустимых названий колонки.
        Значения-заглушки (тире и т. п.) пропускаются."""
        for key in self._CASE_NUMBER_HEADERS:
            v = fields.get(key)
            if v is not None:
                v_str = str(v).strip()
                if v_str and v_str not in self._PLACEHOLDER_VALUES:
                    return v_str
        return ""

    def _resolve_archive_row(self, ws_archive, project: Project) -> int | None:  # noqa: ANN001
        """Определяет номер строки проекта в листе архива.

        Стратегии (в порядке приоритета):
        1. Поиск по номеру дела (если задан и не заглушка).
        2. Использование project.row_index, если он валиден.
        3. Полный перебор строк — выбирается та, у которой больше всего
           совпадающих с project.fields значений ячеек.
        """
        archive_max = ws_archive.max_row or 0
        if archive_max < 2:
            return None

        # 1. По номеру дела
        case_number = self._extract_case_number(project.fields)
        if case_number:
            found = self._find_row_by_case_number(ws_archive, case_number)
            if found is not None and 2 <= found <= archive_max:
                return found

        # 2. По row_index
        if project.row_index is not None and 2 <= project.row_index <= archive_max:
            return project.row_index

        # 3. Перебор строк по максимальному совпадению полей
        headers = [
            str(c.value).strip() if c.value is not None else ""
            for c in ws_archive[1]
        ]
        best_row: int | None = None
        best_score = 0
        for r in range(2, archive_max + 1):
            score = 0
            for col_idx, header in enumerate(headers, start=1):
                if not header:
                    continue
                field_val = str(project.fields.get(header, "") or "").strip()
                if not field_val or field_val in self._PLACEHOLDER_VALUES:
                    continue
                cell_val = ws_archive.cell(row=r, column=col_idx).value
                cell_str = str(cell_val).strip() if cell_val is not None else ""
                if cell_str == field_val:
                    score += 1
            if score > best_score:
                best_score = score
                best_row = r
        return best_row if best_score >= 1 else None

    def _find_row_by_case_number(self, ws, case_number: str) -> int | None:  # noqa: ANN001
        """Возвращает 1-based номер строки, где в колонке с номером дела стоит case_number."""
        case_number = (case_number or "").strip()
        if not case_number:
            return None

        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        col_idx0: int | None = None
        for i, h in enumerate(headers):
            if h.strip().lower() in {n.strip().lower() for n in self._CASE_NUMBER_HEADERS}:
                col_idx0 = i
                break
        if col_idx0 is None:
            return None
        col = col_idx0 + 1  # 1-based

        for r in range(2, (ws.max_row or 0) + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None and str(v).strip() == case_number:
                return r
        return None

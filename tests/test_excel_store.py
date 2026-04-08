"""Тесты загрузки проектов из Excel."""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from filldoc.excel.excel_store import ExcelProjectStore


def test_skips_header_like_rows_in_archive_sheet(tmp_path: Path) -> None:
    """Дубли старой шапки архива не должны попадать в список проектов."""
    excel_path = tmp_path / "projects.xlsx"

    wb = Workbook()
    ws_current = wb.active
    ws_current.title = "Текущие"
    ws_current.append(["Имя проекта", "Кредитор", "Должник", "Новый столбец"])
    ws_current.append(["Текущий проект", "Банк", "ООО", "x"])

    ws_archive = wb.create_sheet("Архив")
    ws_archive.append(["Имя проекта", "Кредитор", "Должник", "Новый столбец"])
    ws_archive.append(["Имя проекта", "Кредитор", "Должник", ""])
    ws_archive.append(["Архивный проект", "Банк 2", "ООО 2", ""])
    wb.save(excel_path)

    store = ExcelProjectStore(str(excel_path))

    archived = store.load_projects_from_sheet("Архив")

    assert [project.fields["Имя проекта"] for project in archived] == ["Архивный проект"]


def test_repair_archive_headers_updates_old_header_in_place(tmp_path: Path) -> None:
    """Починка архива не должна вставлять вторую шапку поверх старой."""
    excel_path = tmp_path / "projects.xlsx"

    wb = Workbook()
    ws_current = wb.active
    ws_current.title = "Текущие"
    ws_current.append(["Имя проекта", "Кредитор", "Должник", "Новый столбец"])
    ws_current.append(["Текущий проект", "Банк", "ООО", "x"])

    ws_archive = wb.create_sheet("Архив")
    ws_archive.append(["Имя проекта", "Кредитор", "Должник", ""])
    ws_archive.append(["Архивный проект", "Банк 2", "ООО 2", ""])
    wb.save(excel_path)

    store = ExcelProjectStore(str(excel_path))

    repaired = store.repair_archive_headers()
    archived = store.load_projects_from_sheet("Архив")

    assert repaired is True
    assert len(archived) == 1
    assert archived[0].fields["Имя проекта"] == "Архивный проект"

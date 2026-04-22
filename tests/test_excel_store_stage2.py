from __future__ import annotations

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))

from filldoc.core.errors import ExcelError
from filldoc.excel.excel_store import ExcelProjectStore
from filldoc.excel.models import FILLDOC_ID_FIELD, Project


def _make_workbook(
    path: Path,
    headers: list[str | None],
    rows: list[list[str]],
    sheet_name: str = "Current",
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


def test_save_project_fields_updates_exact_row(tmp_path: Path) -> None:
    excel_path = _make_workbook(
        tmp_path / "projects.xlsx",
        ["Name", "Status", "Amount"],
        [["First", "Draft", "10"], ["Second", "Draft", "20"]],
    )
    project = Project(
        project_id="row:3",
        row_index=3,
        fields={"Name": "Second updated", "Status": "Ready", "Ignored": "x"},
    )

    ExcelProjectStore(str(excel_path)).save_project_fields(project)

    ws = load_workbook(excel_path).active
    assert [ws.cell(2, col).value for col in range(1, 4)] == ["First", "Draft", "10"]
    assert [ws.cell(3, col).value for col in range(1, 4)] == ["Second updated", "Ready", "20"]


def test_add_project_appends_new_row(tmp_path: Path) -> None:
    excel_path = _make_workbook(
        tmp_path / "projects.xlsx",
        ["Name", "Status", "Amount"],
        [["First", "Draft", "10"]],
    )
    project = Project(
        project_id="new",
        fields={"Name": "Second", "Status": "Ready", "Amount": "20", "Ignored": "x"},
    )

    ExcelProjectStore(str(excel_path)).add_project(project)

    ws = load_workbook(excel_path).active
    assert project.row_index == 3
    assert project.headers == ["Name", "Status", "Amount", FILLDOC_ID_FIELD]
    assert [ws.cell(3, col).value for col in range(1, 4)] == ["Second", "Ready", "20"]
    assert ws.cell(1, 4).value == FILLDOC_ID_FIELD
    assert ws.cell(3, 4).value == project.internal_id


def test_move_project_to_archive_moves_row(tmp_path: Path) -> None:
    excel_path = _make_workbook(
        tmp_path / "projects.xlsx",
        ["Name", "Status"],
        [["Active", "Draft"], ["To archive", "Done"]],
        sheet_name="Current",
    )
    project = Project(
        project_id="row:3",
        row_index=3,
        fields={"Name": "To archive", "Status": "Done"},
    )

    ExcelProjectStore(str(excel_path)).move_project_to_archive(
        project,
        current_sheet_name="Current",
        archive_sheet_name="Archive",
    )

    wb = load_workbook(excel_path)
    assert wb["Current"].max_row == 2
    assert [wb["Current"].cell(2, col).value for col in range(1, 3)] == ["Active", "Draft"]
    assert [wb["Archive"].cell(2, col).value for col in range(1, 3)] == ["To archive", "Done"]
    assert project.row_index == 2


def test_restore_project_from_archive_moves_row_back(tmp_path: Path) -> None:
    excel_path = _make_workbook(
        tmp_path / "projects.xlsx",
        ["Name", "Status"],
        [["Active", "Draft"]],
        sheet_name="Current",
    )
    wb = load_workbook(excel_path)
    ws_archive = wb.create_sheet("Archive")
    ws_archive.append(["Name", "Status"])
    ws_archive.append(["Archived", "Done"])
    wb.save(excel_path)
    project = Project(
        project_id="row:2",
        row_index=2,
        fields={"Name": "Archived", "Status": "Done"},
    )

    ExcelProjectStore(str(excel_path)).restore_project_from_archive(
        project,
        current_sheet_name="Current",
        archive_sheet_name="Archive",
    )

    wb = load_workbook(excel_path)
    assert [wb["Current"].cell(3, col).value for col in range(1, 3)] == ["Archived", "Done"]
    assert wb["Archive"].max_row == 1
    assert project.row_index == 3


def test_missing_excel_file_raises_user_facing_error(tmp_path: Path) -> None:
    store = ExcelProjectStore(str(tmp_path / "missing.xlsx"))

    with pytest.raises(ExcelError):
        store.load_projects()


def test_load_projects_tolerates_empty_and_strange_headers(tmp_path: Path) -> None:
    excel_path = _make_workbook(
        tmp_path / "projects.xlsx",
        [None, "", "Name", "  Status  "],
        [["ignored", "also ignored", "Project A", "Ready"]],
    )

    projects = ExcelProjectStore(str(excel_path)).load_projects()

    assert len(projects) == 1
    assert projects[0].project_id == "row:2"
    assert projects[0].fields["Name"] == "Project A"
    assert projects[0].fields["Status"] == "Ready"
    assert projects[0].fields[FILLDOC_ID_FIELD] == projects[0].internal_id

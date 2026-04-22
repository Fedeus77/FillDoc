from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from filldoc.excel.excel_store import ExcelProjectStore
from filldoc.excel.models import FILLDOC_ID_FIELD
from filldoc.projects.repository import ProjectConflictError, ProjectRepository


def _make_workbook(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Current"
    ws.append(["Name", "Status"])
    ws.append(["First", "Draft"])
    ws.append(["Second", "Ready"])
    wb.save(path)
    return path


def test_load_projects_persists_filldoc_ids(tmp_path: Path) -> None:
    excel_path = _make_workbook(tmp_path / "projects.xlsx")

    projects = ProjectRepository(str(excel_path)).load_projects()

    wb = load_workbook(excel_path)
    ws = wb.active
    assert ws.cell(1, 3).value == FILLDOC_ID_FIELD
    assert [p.internal_id for p in projects] == [ws.cell(2, 3).value, ws.cell(3, 3).value]
    assert all(p.loaded_snapshot for p in projects)


def test_save_project_uses_filldoc_id_when_row_index_is_stale(tmp_path: Path) -> None:
    excel_path = _make_workbook(tmp_path / "projects.xlsx")
    repository = ProjectRepository(str(excel_path))
    projects = repository.load_projects()
    project = projects[1]
    project.fields["Status"] = "Updated"

    wb = load_workbook(excel_path)
    ws = wb.active
    ws.insert_rows(2)
    ws.cell(2, 1).value = "Inserted"
    ws.cell(2, 2).value = "External"
    ws.cell(2, 3).value = "external-id"
    wb.save(excel_path)

    repository.save_project_fields(project, force=True)

    ws = load_workbook(excel_path).active
    assert ws.cell(4, 1).value == "Second"
    assert ws.cell(4, 2).value == "Updated"
    assert project.row_index == 4


def test_repository_detects_external_change_before_save(tmp_path: Path) -> None:
    excel_path = _make_workbook(tmp_path / "projects.xlsx")
    repository = ProjectRepository(str(excel_path))
    project = repository.load_projects()[0]

    wb = load_workbook(excel_path)
    ws = wb.active
    ws.cell(2, 2).value = "Changed outside"
    wb.save(excel_path)

    project.fields["Status"] = "Changed in app"
    with pytest.raises(ProjectConflictError):
        repository.save_project_fields(project)


def test_excel_store_loads_existing_filldoc_id(tmp_path: Path) -> None:
    excel_path = tmp_path / "projects.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", FILLDOC_ID_FIELD])
    ws.append(["Project", "stable-id"])
    wb.save(excel_path)

    project = ExcelProjectStore(str(excel_path)).load_projects()[0]

    assert project.internal_id == "stable-id"
    assert project.fields[FILLDOC_ID_FIELD] == "stable-id"
